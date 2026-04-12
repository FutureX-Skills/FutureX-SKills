#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票处理脚本
从邮箱获取发票邮件 → 下载附件 → 提取字段 → 生成Excel汇总
支持 QQ邮箱、163邮箱、Gmail、Outlook 等任意 IMAP 邮箱
"""

import imaplib
import email
from email.header import decode_header
import re
import os
import sys
import json
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, unquote

# 第三方库延迟导入：缺失时不在此处崩溃，由 check_dependencies() 给出友好提示
try:
    import requests
except ImportError:
    requests = None
try:
    import pdfplumber
except ImportError:
    pdfplumber = None
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None

def log(msg=""):
    """带flush的打印"""
    print(msg, flush=True)


def check_dependencies():
    """检查依赖，缺失时给出安装提示"""
    missing = []
    try:
        import requests
    except ImportError:
        missing.append(("requests", "pip install requests"))
    try:
        import pdfplumber
    except ImportError:
        missing.append(("pdfplumber", "pip install pdfplumber"))
    try:
        import openpyxl
    except ImportError:
        missing.append(("openpyxl", "pip install openpyxl"))
    try:
        from PIL import Image
    except ImportError:
        missing.append(("Pillow", "pip install Pillow"))

    # Playwright和pyzbar是可选依赖，只警告不阻断
    optional_missing = []
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        optional_missing.append(("playwright", "pip install playwright && playwright install chromium"))
    try:
        from pyzbar.pyzbar import decode
    except ImportError:
        optional_missing.append(("pyzbar", "pip install pyzbar（Windows需额外安装libzbar.dll）"))

    if missing:
        log("错误：缺少必需依赖：")
        for name, cmd in missing:
            log(f"  - {name}: {cmd}")
        sys.exit(1)

    if optional_missing:
        log("提示：以下可选依赖未安装，部分功能受限：")
        for name, cmd in optional_missing:
            log(f"  - {name}: {cmd}")
        if any("playwright" in n for n, _ in optional_missing):
            log("  → 无Playwright：需JS渲染的发票链接将无法自动下载")
        if any("pyzbar" in n for n, _ in optional_missing):
            log("  → 无pyzbar：二维码图片将无法自动解码")
        log()


# ==================== 配置 ====================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
CONFIG_FILE = os.path.join(SKILL_DIR, "config.json")


def load_config():
    """从 config.json 读取邮箱配置"""
    template_file = os.path.join(SKILL_DIR, "config.json.template")
    if not os.path.exists(CONFIG_FILE):
        log(f"错误：找不到配置文件 {CONFIG_FILE}")
        if os.path.exists(template_file):
            log(f"请先复制模板：cp config.json.template config.json")
            log(f"然后填写你的邮箱地址和授权码，详见 SETUP.md")
        else:
            log("请先完成首次配置，参考 SETUP.md")
        sys.exit(1)
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    required = ["email", "password", "imap_server"]
    for key in required:
        val = cfg.get(key, "")
        if not val or not val.strip():
            log(f"错误：config.json 中 {key} 为空，请填写后重试")
            log("提示：password 是邮箱授权码，不是登录密码，获取方式见 SETUP.md")
            sys.exit(1)
    return cfg


_config = None


def get_config():
    global _config
    if _config is None:
        _config = load_config()
    return _config



# 发票相关关键词（用于筛选邮件）
INVOICE_KEYWORDS = [
    "发票", "invoice", "电子发票", "增值税", "fapiao",
    "开票", "税务", "发票通知", "电子票据", "数电发票",
    "全电发票", "专用发票", "普通发票",
]

# 发票附件识别关键词（用于从多个附件中识别发票文件）
INVOICE_ATTACHMENT_KEYWORDS = [
    "发票", "invoice", "fapiao", "税", "tax",
]

# 支持的文件类型
PDF_TYPES = ["application/pdf", "application/x-pdf"]
IMAGE_TYPES = ["image/png", "image/jpeg", "image/jpg", "image/tiff"]

# ==================== 邮件获取 ====================


def connect_imap():
    """连接IMAP服务器"""
    cfg = get_config()
    mail = imaplib.IMAP4_SSL(cfg["imap_server"], cfg.get("imap_port", 993))
    mail.login(cfg["email"], cfg["password"])
    return mail


def decode_mime_header(header_value):
    """解码MIME邮件头"""
    if not header_value:
        return ""
    decoded_parts = decode_header(header_value)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            charset = charset or "utf-8"
            try:
                result.append(part.decode(charset))
            except (UnicodeDecodeError, LookupError):
                result.append(part.decode("utf-8", errors="replace"))
        else:
            result.append(part)
    return "".join(result)


def _imap_date(dt):
    """转换为IMAP日期格式 DD-Mon-YYYY"""
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{dt.day:02d}-{months[dt.month - 1]}-{dt.year}"


def search_invoice_emails_imap(mail, start_date, end_date):
    """用IMAP服务端搜索直接定位发票邮件（按关键词+日期）"""
    mail.select("INBOX")

    since_str = _imap_date(start_date)
    before_str = _imap_date(end_date + timedelta(days=1))

    # 用IMAP SUBJECT搜索关键词，服务端过滤，避免全量fetch
    # QQ邮箱IMAP支持OR和SUBJECT
    subject_keywords = ["发票", "invoice", "电子发票", "增值税", "开票", "数电发票", "税务"]
    all_msg_ids = set()

    for keyword in subject_keywords:
        try:
            # IMAP搜索中文需要编码
            criteria = f'(SINCE {since_str} BEFORE {before_str} SUBJECT "{keyword}")'
            status, messages = mail.search("UTF-8", criteria.encode("utf-8"))
            if status == "OK" and messages[0]:
                ids = messages[0].split()
                all_msg_ids.update(ids)
                if ids:
                    log(f"  关键词 '{keyword}' 命中 {len(ids)} 封")
        except Exception:
            # 部分IMAP服务器不支持UTF-8搜索，回退到简单搜索
            try:
                criteria = f'(SINCE {since_str} BEFORE {before_str} SUBJECT "{keyword}")'
                status, messages = mail.search(None, criteria)
                if status == "OK" and messages[0]:
                    ids = messages[0].split()
                    all_msg_ids.update(ids)
            except Exception:
                pass

    msg_ids = sorted(all_msg_ids)
    log(f"IMAP服务端搜索命中 {len(msg_ids)} 封（去重后）")
    return msg_ids


def is_invoice_email(subject, from_addr, body_text):
    """判断是否为发票邮件"""
    text_to_check = f"{subject} {from_addr} {body_text}".lower()
    for keyword in INVOICE_KEYWORDS:
        if keyword.lower() in text_to_check:
            return True
    return False


def get_email_body(msg):
    """提取邮件正文（纯文本和HTML）"""
    text_body = ""
    html_body = ""

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition", ""))
            # 跳过附件
            if "attachment" in content_disposition:
                continue
            try:
                payload = part.get_payload(decode=True)
                if payload is None:
                    continue
                charset = part.get_content_charset() or "utf-8"
                decoded = payload.decode(charset, errors="replace")
                if content_type == "text/plain":
                    text_body += decoded
                elif content_type == "text/html":
                    html_body += decoded
            except Exception:
                continue
    else:
        try:
            payload = msg.get_payload(decode=True)
            if payload:
                charset = msg.get_content_charset() or "utf-8"
                decoded = payload.decode(charset, errors="replace")
                if msg.get_content_type() == "text/html":
                    html_body = decoded
                else:
                    text_body = decoded
        except Exception:
            pass

    return text_body, html_body


def _parse_email_date(date_str):
    """解析邮件Date头，返回datetime或None"""
    if not date_str:
        return None
    from email.utils import parsedate_to_datetime
    try:
        return parsedate_to_datetime(date_str)
    except Exception:
        return None


def fetch_invoice_emails(mail, msg_ids, start_date=None, end_date=None):
    """获取发票邮件详情，支持客户端日期过滤（IMAP日期不靠谱）"""
    invoice_emails = []

    # 第一步：轻量fetch INTERNALDATE过滤日期，减少全量fetch
    filtered_ids = msg_ids
    if start_date and end_date:
        log(f"  正在按日期过滤 {len(msg_ids)} 封邮件...")
        filtered_ids = []
        for msg_id in msg_ids:
            try:
                status, data = mail.fetch(msg_id, "(INTERNALDATE)")
                if status == "OK" and data[0]:
                    date_match = re.search(r'INTERNALDATE "([^"]+)"', data[0].decode() if isinstance(data[0], bytes) else str(data[0]))
                    if date_match:
                        date_text = date_match.group(1)
                        # INTERNALDATE格式: "01-Apr-2026 10:30:00 +0800"
                        from email.utils import parsedate_to_datetime
                        try:
                            dt = parsedate_to_datetime(date_text)
                        except Exception:
                            # 手动解析 DD-Mon-YYYY HH:MM:SS +ZZZZ
                            from datetime import timezone
                            import time
                            try:
                                t = time.strptime(date_text[:20].strip(), "%d-%b-%Y %H:%M:%S")
                                dt = datetime(*t[:6])
                            except Exception:
                                filtered_ids.append(msg_id)
                                continue
                        msg_date = dt.replace(tzinfo=None).date() if hasattr(dt, 'date') else dt
                        if isinstance(msg_date, datetime):
                            msg_date = msg_date.date()
                        if start_date.date() <= msg_date <= end_date.date():
                            filtered_ids.append(msg_id)
            except Exception:
                # 解析失败就保留，后续再过滤
                filtered_ids.append(msg_id)
        log(f"  日期过滤后剩余 {len(filtered_ids)} 封")

    for i, msg_id in enumerate(filtered_ids):
        status, msg_data = mail.fetch(msg_id, "(RFC822)")
        if status != "OK":
            continue

        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)

        subject = decode_mime_header(msg["Subject"])
        from_addr = decode_mime_header(msg["From"])
        date_str = msg["Date"]

        # 获取邮件正文
        text_body, html_body = get_email_body(msg)

        # 判断是否为发票邮件
        if is_invoice_email(subject, from_addr, text_body + html_body):
            invoice_emails.append({
                "msg_id": msg_id,
                "msg": msg,
                "subject": subject,
                "from": from_addr,
                "date": date_str,
                "text_body": text_body,
                "html_body": html_body,
            })
            log(f"  [发票] {subject} | {from_addr}")
        else:
            if (i + 1) % 10 == 0:
                log(f"  已扫描 {i + 1}/{len(msg_ids)} 封...")

    log(f"\n筛选出 {len(invoice_emails)} 封发票邮件")
    return invoice_emails


# ==================== 附件下载 ====================


def extract_urls_from_html(html_text):
    """从HTML中提取URL，自动反转义HTML实体"""
    from html import unescape
    # 先反转义 &amp; → & 等，否则OSS签名链接会被截断
    clean_html = unescape(html_text)
    # 匹配 href 中的链接
    urls = re.findall(r'href=["\']([^"\']+)["\']', clean_html, re.IGNORECASE)
    # 也匹配纯文本中的URL
    urls += re.findall(r'https?://[^\s<>"\']+', clean_html)
    # 去重
    seen = set()
    unique_urls = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            unique_urls.append(url)
    return unique_urls


def is_invoice_url(url):
    """判断URL是否可能是发票下载链接"""
    url_lower = url.lower()
    # 排除明显不相关的链接
    exclude_patterns = [
        "unsubscribe", "mailto:", "javascript:", "#",
        "privacy", "terms", "help", "about",
        "twitter.com", "facebook.com", "weibo.com",
        ".css", "google-analytics",
        "linktrace.", "triggerdelivery",  # 邮件追踪像素
    ]
    for pattern in exclude_patterns:
        if pattern in url_lower:
            return False

    # 排除平台首页（域名匹配但不是下载链接）
    homepage_patterns = [
        r'https?://[^/]+\.(com|cn|com\.cn)/?$',      # 纯域名首页
        r'https?://[^/]+\.(com|cn|com\.cn)/?#/?$',   # SPA首页
    ]
    for pat in homepage_patterns:
        if re.match(pat, url):
            return False

    # 常见发票平台域名（直接命中）
    invoice_domains = [
        "jss.com.cn",       # 诺诺网发票文件服务器
        "nnfp.",            # 诺诺发票短链接
        "inv-file",         # 发票文件路径
        "oss-cn-",          # 阿里云OSS（常用于发票存储）
    ]
    for domain in invoice_domains:
        if domain in url_lower:
            return True

    # 关键词匹配
    invoice_indicators = [
        "fapiao", "invoice", "发票", "download", "pdf",
        "einvoice", "etax", "tax", "bill",
    ]
    for indicator in invoice_indicators:
        if indicator in url_lower:
            return True

    # 文件扩展名匹配
    if re.search(r'\.(pdf|png|jpg|jpeg)(\?|$)', url_lower):
        return True
    return False


def _try_decode_qr_code(filepath):
    """尝试解码图片中的二维码，返回解码出的URL或None"""
    try:
        from pyzbar.pyzbar import decode
        from PIL import Image
        img = Image.open(filepath)
        results = decode(img)
        for result in results:
            data = result.data.decode("utf-8", errors="ignore")
            if data.startswith("http://") or data.startswith("https://"):
                return data
        return None
    except ImportError:
        return None
    except Exception:
        return None


def _is_valid_invoice_file(filepath):
    """验证下载的文件是否像发票，过滤掉消费明细、logo等非发票文件"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".pdf":
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages[:2]:  # 只看前两页
                    t = page.extract_text()
                    if t:
                        text += t
            if not text:
                return True  # 无法提取文本，保留（可能是扫描件）
            # 发票必须包含至少一个关键特征
            invoice_markers = ["发票号码", "发票代码", "价税合计", "税额", "增值税", "电子发票", "数电发票"]
            return any(m in text for m in invoice_markers)
        except Exception:
            return True  # 解析失败，保留

    elif ext in (".png", ".jpg", ".jpeg"):
        try:
            from PIL import Image
            img = Image.open(filepath)
            w, h = img.size
            # 发票图片至少有一定尺寸（短边>300px），小图标/logo过滤掉
            if min(w, h) < 300:
                return False
            return True
        except Exception:
            return True

    return True


def _save_response_file(content, content_type, url, output_dir, prefix):
    """保存下载的文件内容，返回文件路径或None"""
    # 先通过magic bytes判断真实类型，防止HTML页面被当文件保存
    is_pdf = False
    is_image = False

    if content[:4] == b'%PDF':
        is_pdf = True
    elif content[:8] == b'\x89PNG\r\n\x1a\n':
        is_image = True
    elif content[:3] == b'\xff\xd8\xff':
        is_image = True  # JPEG magic bytes
    elif b'<html' in content[:1024].lower() or b'<!doctype' in content[:1024].lower():
        # 这是HTML页面，不是文件
        return None

    # magic bytes没命中时，才看Content-Type和URL
    if not (is_pdf or is_image):
        is_pdf = "pdf" in content_type or url.lower().endswith(".pdf")
        is_image = any(t in content_type for t in ["image/png", "image/jpeg", "image/jpg"])

    if not (is_pdf or is_image):
        return None

    ext = ".pdf" if is_pdf else ".png" if is_image and b'\x89PNG' in content[:8] else ".jpg"
    filename = f"{prefix}{ext}"
    filepath = os.path.join(output_dir, filename)

    with open(filepath, "wb") as f:
        f.write(content)

    size_kb = len(content) / 1024
    if size_kb < 5:
        os.remove(filepath)
        return None

    # 验证是否为发票文件
    if not _is_valid_invoice_file(filepath):
        log(f"    跳过非发票文件: {filename}")
        os.remove(filepath)
        return None

    log(f"    下载成功: {filename} ({size_kb:.1f}KB)")
    return filepath


def try_download_url(url, output_dir, prefix, log_entries):
    """尝试通过HTTP下载，失败则用Playwright浏览器回退"""
    # 第一步：HTTP直接下载
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/120.0.0.0 Safari/537.36",
        }
        resp = requests.get(url, headers=headers, timeout=30,
                            allow_redirects=True, stream=False)
        resp.raise_for_status()

        content_type = resp.headers.get("Content-Type", "").lower()

        # 如果是直接文件，保存
        if "html" not in content_type:
            result = _save_response_file(resp.content, content_type, resp.url,
                                         output_dir, prefix)
            if result:
                return result

        # 如果是HTML页面，尝试Playwright回退
        log(f"    HTTP返回HTML页面，尝试浏览器下载...")
        return _try_playwright_download(url, output_dir, prefix, log_entries)

    except Exception as e:
        # HTTP失败也尝试Playwright
        log(f"    HTTP下载失败({str(e)[:50]})，尝试浏览器下载...")
        return _try_playwright_download(url, output_dir, prefix, log_entries)


def _try_playwright_download(url, output_dir, prefix, log_entries):
    """用Playwright打开链接，捕获下载或打印为PDF"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log_entries.append(f"[需浏览器] URL: {url} | Playwright未安装")
        return None

    def _save_download(download, output_dir, prefix):
        """从Playwright Download对象保存文件"""
        ext = ".pdf"
        fname = download.suggested_filename or ""
        if fname.lower().endswith((".png", ".jpg", ".jpeg")):
            ext = os.path.splitext(fname)[1]
        filepath = os.path.join(output_dir, f"{prefix}{ext}")
        download.save_as(filepath)
        size_kb = os.path.getsize(filepath) / 1024
        if size_kb < 5:
            os.remove(filepath)
            return None
        return filepath

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()

            # 方法1：goto直接触发下载（OSS直链等）
            try:
                with page.expect_download(timeout=15000) as dl_info:
                    page.goto(url, wait_until="commit", timeout=30000)
                download = dl_info.value
                filepath = _save_download(download, output_dir, prefix)
                if filepath:
                    log(f"    浏览器直接下载: {os.path.basename(filepath)} ({os.path.getsize(filepath)/1024:.1f}KB)")
                    browser.close()
                    return filepath
            except Exception:
                # 没触发下载，继续等页面加载
                try:
                    page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass

            # 方法2：尝试点击页面上的下载按钮
            download_selectors = [
                'a:has-text("下载")', 'button:has-text("下载")',
                'a:has-text("PDF")', 'button:has-text("PDF")',
                'a[download]', '[class*="download"]',
            ]
            for selector in download_selectors:
                try:
                    el = page.query_selector(selector)
                    if el and el.is_visible():
                        with page.expect_download(timeout=15000) as dl_info:
                            el.click()
                        download = dl_info.value
                        ext = ".pdf"
                        fname = download.suggested_filename or ""
                        if fname.lower().endswith((".png", ".jpg", ".jpeg")):
                            ext = os.path.splitext(fname)[1]
                        filepath = os.path.join(output_dir, f"{prefix}{ext}")
                        download.save_as(filepath)
                        size_kb = os.path.getsize(filepath) / 1024
                        log(f"    浏览器点击下载成功: {os.path.basename(filepath)} ({size_kb:.1f}KB)")
                        browser.close()
                        return filepath
                except Exception:
                    continue

            # 方法3：页面可能直接展示了发票（如内嵌PDF），用打印导出
            # 检查页面是否有embed/iframe包含PDF
            pdf_embed = page.query_selector('embed[type="application/pdf"], iframe[src*=".pdf"]')
            if pdf_embed:
                src = pdf_embed.get_attribute("src")
                if src:
                    # 直接下载嵌入的PDF
                    resp = page.request.get(src)
                    if resp.ok:
                        filepath = os.path.join(output_dir, f"{prefix}.pdf")
                        with open(filepath, "wb") as f:
                            f.write(resp.body())
                        size_kb = len(resp.body()) / 1024
                        log(f"    从嵌入PDF提取: {os.path.basename(filepath)} ({size_kb:.1f}KB)")
                        browser.close()
                        return filepath

            browser.close()
            log_entries.append(f"[浏览器未找到下载] URL: {url}")
            return None

    except Exception as e:
        log_entries.append(f"[浏览器下载失败] URL: {url} | 错误: {str(e)[:100]}")
        return None


def download_attachments(invoice_email, output_dir, index, log_entries):
    """从一封发票邮件中下载附件"""
    msg = invoice_email["msg"]
    subject = invoice_email["subject"]
    downloaded = []

    # 第一步：提取直接附件
    attachment_idx = 0
    for part in msg.walk():
        content_disposition = str(part.get("Content-Disposition", ""))
        if "attachment" not in content_disposition and "inline" not in content_disposition:
            continue

        filename = part.get_filename()
        if filename:
            filename = decode_mime_header(filename)
        content_type = part.get_content_type()

        # 判断是否为PDF或图片
        is_pdf = (content_type in PDF_TYPES or
                  (filename and filename.lower().endswith(".pdf")))
        is_image = (content_type in IMAGE_TYPES or
                    (filename and filename.lower().endswith((".png", ".jpg", ".jpeg"))))

        if not (is_pdf or is_image):
            continue

        # 如果有多个附件，优先PDF
        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        attachment_idx += 1
        ext = ".pdf" if is_pdf else os.path.splitext(filename)[1] if filename else ".jpg"
        safe_subject = re.sub(r'[\\/:*?"<>|]', '', subject)[:30]
        # 多附件加编号防覆盖
        if attachment_idx == 1:
            out_filename = f"{index:03d}-{safe_subject}{ext}"
        else:
            out_filename = f"{index:03d}-{safe_subject}-{attachment_idx}{ext}"
        filepath = os.path.join(output_dir, out_filename)

        with open(filepath, "wb") as f:
            f.write(payload)

        # 验证是否为发票文件（过滤消费明细、logo等）
        if not _is_valid_invoice_file(filepath):
            log(f"    跳过非发票附件: {out_filename}")
            os.remove(filepath)
            continue

        # 图片附件：检查是否为二维码，如果是则尝试下载真正的发票
        if is_image:
            qr_url = _try_decode_qr_code(filepath)
            if qr_url:
                log(f"    附件是二维码，解码链接: {qr_url[:80]}...")
                qr_prefix = f"{index:03d}-qr-invoice"
                qr_result = try_download_url(qr_url, output_dir, qr_prefix, log_entries)
                if qr_result:
                    log(f"    二维码链路成功: {os.path.basename(qr_result)}")
                    os.remove(filepath)  # 删除原二维码图片
                    downloaded.append(qr_result)
                    continue
                else:
                    log(f"    二维码链路未获取到发票，保留原图片")

        size_kb = len(payload) / 1024
        log(f"    附件: {out_filename} ({size_kb:.1f}KB)")
        downloaded.append(filepath)

    # 第二步：如果没有直接附件，尝试从正文提取链接下载
    if not downloaded:
        html_body = invoice_email.get("html_body", "")
        text_body = invoice_email.get("text_body", "")
        all_text = html_body + text_body

        urls = extract_urls_from_html(all_text)
        invoice_urls = [u for u in urls if is_invoice_url(u)]

        if invoice_urls:
            log(f"    找到 {len(invoice_urls)} 个可能的发票链接，尝试下载...")
            for url_idx, url in enumerate(invoice_urls[:5]):  # 限制最多试5个链接
                safe_subject = re.sub(r'[\\/:*?"<>|]', '', subject)[:30]
                prefix = f"{index:03d}-{safe_subject}-link{url_idx + 1}"
                result = try_download_url(url, output_dir, prefix, log_entries)
                if result:
                    downloaded.append(result)
                    break  # 下到一个就够了

        if not downloaded:
            # 记录需要手动处理的
            if invoice_urls:
                log_entries.append(
                    f"[需手动下载] 邮件: {subject} | "
                    f"链接: {invoice_urls[0]}"
                )
            else:
                log_entries.append(
                    f"[无附件无链接] 邮件: {subject} | 发件人: {invoice_email['from']}"
                )

    return downloaded


# ==================== 发票字段提取 ====================


def extract_fields_from_table(tables, filepath):
    """从pdfplumber表格中提取发票字段（主策略）"""
    fields = {
        "文件": os.path.basename(filepath),
        "发票号码": "", "发票代码": "", "开票日期": "",
        "销售方": "", "购买方": "",
        "金额": "", "税额": "", "价税合计": "",
    }

    for table in tables:
        for row in table:
            if not row:
                continue
            cells = [str(c).replace('\n', ' ').strip() if c else '' for c in row]
            row_text = ' '.join(cells)

            # 购买方/销售方：找包含"名称："的单元格
            for cell_text in cells:
                if not cell_text:
                    continue
                # 从"名称：XXX 统一..."中提取公司名
                m = re.search(r'名称[：:]\s*(.+?)(?:\s+统一|$)', cell_text)
                if m:
                    name = m.group(1).strip()
                    if len(name) < 2:
                        continue
                    # 判断是购买方还是销售方：看同行是否有"购"或"销"标记
                    cell_idx = cells.index(cell_text) if cell_text in cells else -1
                    # 检查前面的单元格是否含"购"/"销"
                    is_buyer = False
                    is_seller = False
                    for prev_cell in cells[:max(cell_idx, 0) + 1]:
                        if re.search(r'购\s*买?\s*方', prev_cell):
                            is_buyer = True
                        elif re.search(r'销\s*售?\s*方', prev_cell):
                            is_seller = True
                    # 也通过位置判断：靠左的通常是购买方
                    if not is_buyer and not is_seller:
                        if cell_idx <= len(cells) // 2:
                            is_buyer = True
                        else:
                            is_seller = True
                    if is_buyer and not fields["购买方"]:
                        fields["购买方"] = name
                    elif is_seller and not fields["销售方"]:
                        fields["销售方"] = name

            # 价税合计
            if '价税合计' in row_text and not fields["价税合计"]:
                m = re.search(r'[（(]?\s*小写\s*[)）]?\s*[¥￥]\s*([\d,]+\.?\d*)', row_text)
                if m:
                    fields["价税合计"] = m.group(1).replace(",", "")

    # 检查是否提取到了核心字段
    if fields["购买方"] or fields["销售方"] or fields["价税合计"]:
        return fields
    return None


def extract_fields_from_pdf(filepath):
    """从PDF发票中提取字段：表格提取优先，文本正则兜底"""
    try:
        with pdfplumber.open(filepath) as pdf:
            all_tables = []
            full_text = ""
            for page in pdf.pages:
                tables = page.extract_tables()
                all_tables.extend(tables)
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

        # 策略1：表格提取（更可靠）
        if all_tables:
            fields = extract_fields_from_table(all_tables, filepath)
            if fields:
                # 表格拿不到发票号码和日期，从文本补充
                if full_text.strip():
                    text_fields = parse_invoice_text(full_text, filepath)
                    if text_fields:
                        for key in ["发票号码", "发票代码", "开票日期", "金额", "税额"]:
                            if not fields.get(key) and text_fields.get(key):
                                fields[key] = text_fields[key]
                return fields

        # 策略2：文本正则兜底
        if full_text.strip():
            return parse_invoice_text(full_text, filepath)

        return None

    except Exception as e:
        return {"_error": f"PDF解析失败: {str(e)}", "_file": filepath}


def parse_invoice_text(text, filepath):
    """从发票文本中正则提取字段"""
    fields = {
        "文件": os.path.basename(filepath),
        "发票号码": "",
        "发票代码": "",
        "开票日期": "",
        "销售方": "",
        "购买方": "",
        "金额": "",
        "税额": "",
        "价税合计": "",
    }

    # 发票号码（通常8位或20位数字）
    m = re.search(r'发票号码[：:\s]*(\d{8,20})', text)
    if m:
        fields["发票号码"] = m.group(1)
    else:
        # 数电发票号码可能是20位
        m = re.search(r'[Nn][Oo][.．:：\s]*(\d{8,20})', text)
        if m:
            fields["发票号码"] = m.group(1)
        else:
            # 回退：表格式PDF中号码和标签分离，找独立的20位数字（排除信用代码行）
            lines = text.split('\n')
            for line in lines:
                line_stripped = line.strip()
                # 跳过信用代码行（包含字母）
                if re.search(r'[A-Za-z]', line_stripped):
                    continue
                m = re.search(r'(\d{20})', line_stripped)
                if m:
                    fields["发票号码"] = m.group(1)
                    break

    # 发票代码（10-12位数字，数电发票可能没有）
    m = re.search(r'发票代码[：:\s]*(\d{10,12})', text)
    if m:
        fields["发票代码"] = m.group(1)

    # 开票日期
    m = re.search(r'开票日期[：:\s]*([\d]{4})\s*年\s*([\d]{1,2})\s*月\s*([\d]{1,2})\s*日', text)
    if m:
        fields["开票日期"] = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    else:
        # 回退：匹配任意位置的 YYYY年MM月DD日
        m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', text)
        if m:
            fields["开票日期"] = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
        else:
            m = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', text)
            if m:
                fields["开票日期"] = m.group(1)

    # 用于截断公司名称后面跟着的字段标签
    _name_stop = r'(?:纳税人识别号|统一社会信用代码|统一信用代码|地址|电话|开户行|账号|销\s*名称|购\s*名称)'

    # 购买方名称（兼容数电发票 "购 名称：" 格式）
    buyer_patterns = [
        r'(?:购买方|购方|受票方)[名称]*[：:\s]*([^\n]{2,50})',
        r'购\s+名称[：:\s]*([^\n]{2,50})',
    ]
    for pat in buyer_patterns:
        m = re.search(pat, text)
        if m:
            name = m.group(1).strip()
            name = re.split(_name_stop, name)[0].strip()
            if len(name) >= 2:
                fields["购买方"] = name
                break

    # 销售方名称（兼容数电发票 "销 名称：" 格式）
    seller_patterns = [
        r'(?:销售方|销方|开票方)[名称]*[：:\s]*([^\n]{2,50})',
        r'销\s+名称[：:\s]*([^\n]{2,50})',
    ]
    for pat in seller_patterns:
        m = re.search(pat, text)
        if m:
            name = m.group(1).strip()
            name = re.split(_name_stop, name)[0].strip()
            if len(name) >= 2:
                fields["销售方"] = name
                break

    # 回退策略：表格式布局中标签和值分离，通过信用代码行定位公司名
    # pdfplumber提取表格PDF时，"名称："后面为空，公司名出现在独立行中
    # 典型模式：一行两个公司名，下一行两个信用代码
    if not fields["购买方"] or not fields["销售方"]:
        lines = text.split('\n')
        for i, line in enumerate(lines):
            # 找包含两个信用代码的行（18位，数字字母混合）
            codes = re.findall(r'[0-9A-Z]{15,20}', line)
            if len(codes) >= 2:
                # 上一行应该包含两个公司名
                if i > 0:
                    prev_line = lines[i - 1].strip()
                    # 按公司名常见后缀切分：有限公司、餐厅、酒店 等
                    companies = re.findall(r'[\u4e00-\u9fa5（()）]+(?:有限公司|餐厅|酒店|门店|商行|工作室|事务所|中心|餐馆|饭店)', prev_line)
                    if len(companies) >= 2:
                        if not fields["购买方"]:
                            fields["购买方"] = companies[0]
                        if not fields["销售方"]:
                            fields["销售方"] = companies[1]
                break

    # 价税合计（大写金额前面通常有小写金额）
    m = re.search(r'价税合计[（(小写)）]*[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)', text)
    if m:
        fields["价税合计"] = m.group(1).replace(",", "")
    else:
        # 也尝试匹配"小写"后面的金额
        m = re.search(r'小写[）)]*[：:\s]*[¥￥]?\s*([\d,]+\.\d{2})', text)
        if m:
            fields["价税合计"] = m.group(1).replace(",", "")

    # 金额（不含税）
    m = re.search(r'合\s*计[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)', text)
    if m:
        fields["金额"] = m.group(1).replace(",", "")

    # 税额
    m = re.search(r'税\s*额[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)', text)
    if m:
        fields["税额"] = m.group(1).replace(",", "")

    # 如果有价税合计但缺金额或税额，尝试推算
    if fields["价税合计"] and fields["金额"] and not fields["税额"]:
        try:
            total = float(fields["价税合计"])
            amount = float(fields["金额"])
            fields["税额"] = f"{total - amount:.2f}"
        except ValueError:
            pass

    return fields


def extract_fields_from_image(filepath, output_dir=None, log_entries=None):
    """图片发票 — 尝试二维码解码+浏览器下载，失败则标记为待AI识别"""
    # 尝试解码二维码
    qr_url = _try_decode_qr_code(filepath)
    if qr_url and output_dir:
        log(f"    检测到二维码链接: {qr_url[:80]}...")
        prefix = os.path.splitext(os.path.basename(filepath))[0] + "-qr"
        result = try_download_url(qr_url, output_dir, prefix, log_entries or [])
        if result and result.lower().endswith(".pdf"):
            # 下载到PDF了，递归提取字段
            return extract_fields_from_pdf(result)
        elif result:
            log(f"    二维码下载了非PDF文件: {os.path.basename(result)}")

    return {
        "文件": os.path.basename(filepath),
        "发票号码": "",
        "发票代码": "",
        "开票日期": "",
        "销售方": "",
        "购买方": "",
        "金额": "",
        "税额": "",
        "价税合计": "",
        "_note": "图片发票，需AI视觉识别",
    }


def extract_invoice_fields(filepath, output_dir=None, log_entries=None):
    """根据文件类型提取发票字段"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".pdf":
        return extract_fields_from_pdf(filepath)
    elif ext in (".png", ".jpg", ".jpeg"):
        return extract_fields_from_image(filepath, output_dir, log_entries)
    else:
        return None


# ==================== Excel 生成 ====================


def generate_excel(all_invoice_data, output_dir):
    """按购买方分sheet生成Excel"""
    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 按购买方分组
    grouped = {}
    unclassified = []
    for data in all_invoice_data:
        buyer = data.get("购买方", "").strip()
        if not buyer:
            buyer = "未识别购买方"
        if buyer not in grouped:
            grouped[buyer] = []
        grouped[buyer].append(data)

    if not grouped:
        grouped["无数据"] = []

    headers = ["文件", "发票号码", "发票代码", "开票日期", "销售方",
               "购买方", "金额", "税额", "价税合计", "备注"]

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for buyer, invoices in sorted(grouped.items()):
        # sheet名最长31字符，去掉非法字符
        sheet_name = re.sub(r'[\\/:*?\[\]]', '', buyer)[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 写表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # 写数据行
        for row_idx, data in enumerate(invoices, 2):
            note = data.get("_note", "") or data.get("_error", "")
            row_data = [
                data.get("文件", ""),
                data.get("发票号码", ""),
                data.get("发票代码", ""),
                data.get("开票日期", ""),
                data.get("销售方", ""),
                data.get("购买方", ""),
                data.get("金额", ""),
                data.get("税额", ""),
                data.get("价税合计", ""),
                note,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # 调整列宽
        col_widths = [30, 22, 14, 14, 25, 25, 14, 14, 14, 25]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 合计行
        if invoices:
            total_row = len(invoices) + 2
            ws.cell(row=total_row, column=5, value="合计").font = Font(bold=True)
            for col_idx, col_name in [(7, "金额"), (8, "税额"), (9, "价税合计")]:
                total = 0
                for data in invoices:
                    try:
                        total += float(data.get(col_name, 0) or 0)
                    except (ValueError, TypeError):
                        pass
                if total > 0:
                    cell = ws.cell(row=total_row, column=col_idx, value=round(total, 2))
                    cell.font = Font(bold=True)
                    cell.border = thin_border

    # 保存
    excel_path = os.path.join(output_dir, "发票汇总.xlsx")
    wb.save(excel_path)
    log(f"\nExcel已生成: {excel_path}")
    return excel_path


# ==================== 主流程 ====================


def parse_date_range(date_arg):
    """解析日期参数，支持多种格式"""
    # 格式1: 2026-03（整月）
    m = re.match(r'^(\d{4})-(\d{1,2})$', date_arg)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        start = datetime(year, month, 1)
        # 计算月末
        if month == 12:
            end = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end = datetime(year, month + 1, 1) - timedelta(days=1)
        return start, end

    # 格式2: 2026-03-01~2026-03-15（日期范围）
    m = re.match(r'^(\d{4}-\d{1,2}-\d{1,2})[~\-至到](\d{4}-\d{1,2}-\d{1,2})$', date_arg)
    if m:
        start = datetime.strptime(m.group(1), "%Y-%m-%d")
        end = datetime.strptime(m.group(2), "%Y-%m-%d")
        return start, end

    # 格式3: 2026-03-15（单日）
    m = re.match(r'^(\d{4}-\d{1,2}-\d{1,2})$', date_arg)
    if m:
        d = datetime.strptime(m.group(1), "%Y-%m-%d")
        return d, d

    raise ValueError(f"无法解析日期: {date_arg}，支持格式: 2026-03 / 2026-03-01~2026-03-15 / 2026-03-15")


def main():
    parser = argparse.ArgumentParser(description="邮箱发票处理工具")
    parser.add_argument("date", help="日期范围: 2026-03(整月) / 2026-03-01~2026-03-15(范围) / 2026-03-15(单日)")
    parser.add_argument("--output", "-o", help="输出目录（默认桌面）", default=None)
    args = parser.parse_args()

    # 检查依赖
    check_dependencies()

    # 解析日期
    start_date, end_date = parse_date_range(args.date)
    log(f"日期范围: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")

    # 创建输出目录（跨平台：优先用系统桌面路径）
    if args.output:
        output_dir = args.output
    else:
        date_label = args.date.replace("~", "-to-")
        if sys.platform == "win32":
            desktop = os.path.join(os.environ.get("USERPROFILE", os.path.expanduser("~")), "Desktop")
        else:
            desktop = os.path.expanduser("~/Desktop")
        output_dir = os.path.join(desktop, f"发票-{date_label}")

    os.makedirs(output_dir, exist_ok=True)
    log(f"输出目录: {output_dir}")

    log_entries = []

    # 步骤1: 连接邮箱
    cfg = get_config()
    log(f"\n=== 步骤1: 连接邮箱 ({cfg['email']}) ===")
    mail = connect_imap()
    log("连接成功")

    # 步骤2: IMAP服务端搜索发票邮件（按关键词+日期，避免全量fetch）
    log("\n=== 步骤2: 搜索发票邮件 ===")
    msg_ids = search_invoice_emails_imap(mail, start_date, end_date)

    if not msg_ids:
        log("该日期范围内没有发票相关邮件")
        mail.logout()
        return

    # 步骤3: 获取邮件详情
    log("\n=== 步骤3: 获取邮件详情 ===")
    invoice_emails = fetch_invoice_emails(mail, msg_ids, start_date, end_date)
    mail.logout()

    if not invoice_emails:
        log("没有找到发票邮件")
        return

    # 步骤4: 下载附件
    log("\n=== 步骤4: 下载发票附件 ===")
    all_files = []
    for i, inv_email in enumerate(invoice_emails, 1):
        log(f"\n  [{i}/{len(invoice_emails)}] {inv_email['subject']}")
        files = download_attachments(inv_email, output_dir, i, log_entries)
        all_files.extend(files)

    log(f"\n共下载 {len(all_files)} 个文件")

    # 步骤5: 提取发票字段
    log("\n=== 步骤5: 提取发票字段 ===")
    all_invoice_data = []
    for filepath in all_files:
        log(f"  解析: {os.path.basename(filepath)}")
        fields = extract_invoice_fields(filepath, output_dir, log_entries)
        if fields:
            all_invoice_data.append(fields)
            buyer = fields.get("购买方", "未识别")
            seller = fields.get("销售方", "未识别")
            total = fields.get("价税合计", "未识别")
            log(f"    购买方: {buyer} | 销售方: {seller} | 价税合计: {total}")
        else:
            log_entries.append(f"[解析失败] 文件: {os.path.basename(filepath)}")

    # 步骤6: 生成Excel
    log("\n=== 步骤6: 生成Excel ===")
    if all_invoice_data:
        generate_excel(all_invoice_data, output_dir)
    else:
        log("没有成功解析的发票数据，跳过Excel生成")

    # 步骤7: 写日志
    log_path = os.path.join(output_dir, "处理日志.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"发票处理日志\n")
        f.write(f"处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"日期范围: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}\n")
        f.write(f"IMAP命中: {len(msg_ids)} 封\n")
        f.write(f"发票邮件: {len(invoice_emails)} 封\n")
        f.write(f"下载文件: {len(all_files)} 个\n")
        f.write(f"解析成功: {len(all_invoice_data)} 个\n")
        f.write(f"\n{'='*60}\n\n")

        if log_entries:
            f.write("需要关注的事项:\n\n")
            for entry in log_entries:
                f.write(f"  {entry}\n")
        else:
            f.write("所有发票处理成功，无异常。\n")

    log(f"日志已保存: {log_path}")

    # 打印摘要
    log(f"\n{'='*60}")
    log(f"处理完成!")
    log(f"  发票邮件: {len(invoice_emails)} 封")
    log(f"  下载文件: {len(all_files)} 个")
    log(f"  解析成功: {len(all_invoice_data)} 个")
    if log_entries:
        log(f"  需关注: {len(log_entries)} 项（见处理日志.txt）")
    log(f"  输出目录: {output_dir}")


if __name__ == "__main__":
    main()
